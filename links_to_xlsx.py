# Importa funções de web_scrapping
from asyncio.subprocess import Process
from asyncio.windows_events import NULL
from multiprocessing import process
import web_scrapping
# Importa biblioteca que lida com funções de criação do xlsx
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
# Importa requests_html para download do conteúdo da página
from requests_html import HTMLSession
# Importa os para remoção de arquivo de resolução
import os
# Adiciona barra de progresso às requisições e webscrappings
from progress.bar import ShadyBar

# Função que salva texto da resolução em arquivo


def salvarResolu(text, path, number) -> str:
    with open(f'{path}/resolucao_{str(number)}.txt', 'w+', encoding='utf-8') as file:
        file.write(text)
    with open(f'{path}/resolucao_{str(number)}.txt', 'r', encoding='utf-8') as file:
        resolu = file.readline()
    os.remove(f'{path}/resolucao_{str(number)}.txt')
    with open(f'{path}/' + str(resolu).strip('\n') + '.txt', 'w+', encoding='utf-8') as file:
        file.write(text.split('ANEXO', 1)[1] + '\n')
    return resolu

# Função que salva estado atual de xlsx


def salvarExcel(path) -> None:
    wb.save(path)


# Verifica se os sites de petições estão acessíveis
web_scrapping.siteAcessivel(web_scrapping.lerResolucoes())
# Realiza o download da página de petições
web_scrapping.downloadDoLinkDaPagina()
# Recebe lista de páginas a serem baixadas e manipuladas
list = web_scrapping.listaDePaginas()
# Cria Diretório para armazenar arquivos das páginas
caminho = web_scrapping.criarDiretorio()
# Limpa a tela
web_scrapping.limparTela()
# Inicializa a barra de progresso
bar = ShadyBar('Processando os dados:', max=len(list))
# Atribui cada resolução à uma variável
for i in range(len(list)):
    session = HTMLSession()
    link = session.get(list[i])
    # Resolve conteúdo acessível por javascript
    link.html.render()  # type: ignore
    # Atribui todo condeúdo de elemento com dados da resolução à variável resolucoes
    resolucoes = link.html.find(  # type: ignore
        "#materia > div > div.texto-dou", first=True).text  # type: ignore
    # Atribui diretório do query a uma variável
    path = web_scrapping.criarDiretorio()
    # Salva resolução completa em arquivo de texto e retorna nome da resolução
    resolu = salvarResolu(resolucoes, path, i)
    # Abre novo workbook para adicionar os dados da resolução
    wb = openpyxl.Workbook()
    # Define nome do excel para dada resolução
    wbname = f'{path}/' + str(resolu).strip('\n') + '.xlsx'
    # Cria arquivo excel para dada resolução
    salvarExcel(wbname)
    wb.close()
    # Cria uma versão ativa do worksheet a ser trabalhado
    wb = openpyxl.load_workbook(f'{path}/' + str(resolu).strip('\n') + '.xlsx')
    wb_active = wb.active
    # Define os dados a serem inseridos na primeira coluna de cada excel:
    dados = {1: "RESOLUÇÃO", 2: "EMPRESA", 3: "AUTORIZAÇÃO", 4: "MARCA", 5: "PROCESSO",
             6: "REGISTRO", 7: "VENDA E EMPREGO", 8: "VENCIMENTO", 9: "APRESENTAÇÃO",
             10: "VALIDADE", 11: "CATEGORIA", 12: "ASSUNTO\nPETIÇÃO", 13: "EXPEDIENTE E PETIÇÃO",
             14: "VERSÃO"}

    wb_active.insert_rows(200)
    for column, value in dados.items():
        wb_active.cell(row=1, column=column, value=value)
    # Percorre toas as linhas do arquivo da resolucao e insere na tabela
    with open(f'{path}/' + resolu.strip('\n') + '.txt', 'r+', encoding='utf-8') as file:
        k = 2
        registro = None
        for line in file:
            if "NOME DA EMPRESA:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=1, value=resolu)
                empresa = line.split(":", 1)[1].rstrip('\n') if (line.split(
                    ":", 1)[1].rstrip('\n')) != NULL else line.split(":", 1)[1].rstrip('\n')
            elif "AUTORIZAÇÃO:" in line.rstrip('\n'):
                autorizacao = line.split(":", 1)[1].rstrip('\n') if (line.split(
                    ":", 1)[1].rstrip('\n')) != NULL else line.split(":", 1)[1].rstrip('\n')
            elif "NOME DO PRODUTO E MARCA:" in line.rstrip('\n'):
                produto = line.split(":", 1)[1].rstrip('\n') if (line.split(
                    ":", 1)[1].rstrip('\n')) != NULL else line.split(":", 1)[1].rstrip('\n')
            elif "NUMERO DE PROCESSO:" in line.rstrip('\n'):
                processo = line.split(":", 1)[1].rstrip('\n') if (line.split(
                    ":", 1)[1].rstrip('\n')) != NULL else line.split(":", 1)[1].rstrip('\n')
            elif "NUMERO DE REGISTRO:" in line.rstrip('\n'):
                k += 1
                wb_active.cell(row=k, column=4, value=produto)  # type: ignore
                # type: ignore
                wb_active.cell(row=k, column=3, value=autorizacao)# type: ignore
                wb_active.cell(row=k, column=2, value=str(
                    empresa))  # type: ignore
                wb_active.cell(row=k, column=5, value=processo)  # type: ignore
                wb_active.cell(row=k, column=6, value=line.split(
                    ":", 1)[1].rstrip('\n'))
                wb_active.cell(row=k, column=1, value=resolu)

            elif "VENDA E EMPREGO:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=7, value=line.split(
                    ":", 1)[1].rstrip('\n'))
            elif "VENCIMENTO:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=8, value=line.split(
                    ":", 1)[1].rstrip('\n'))
            elif "APRESENTAÇÃO:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=9, value=line.split(
                    ":", 1)[1].rstrip('\n'))
            elif "VALIDADE DO PRODUTO:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=10, value=line.split(
                    ":", 1)[1].rstrip('\n'))
            elif "CATEGORIA:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=11, value=line.split(
                    ":", 1)[1].rstrip('\n'))
            elif "ASSUNTO DA PETIÇÃO:" in line.rstrip('\n'):
                wb_active.cell(row=k, column=12, value=line.split(
                    ":", 1)[1].rstrip('\n'))
        salvarExcel(wbname)
    bar.next()
    # Fecha wb sendo trabalhado para possibilitar abrir novo arquivo excel
    wb.close()
bar.finish()
